import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import random
import json
import openpyxl
import pandas



headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36"

}

def get_source_html(url):
    driver = webdriver.Chrome(
        executable_path='chromedriver/chromedriver.exe'
    )
    driver.maximize_window()

    try:
        driver.get(url=url)
        time.sleep(3)
        count = 1

        with open(f'pages_test/page{count}.html', 'w', encoding='utf-8') as file:
            file.write(driver.page_source)
        print(f"Processing page {count}")
        count += 1
        while True:
            driver.get(url+f'?page={count}')
            if driver.current_url == url:
                print("End of pages")
                break
            with open(f'pages_test/page{count}.html', 'w', encoding='utf-8') as file:
                file.write(driver.page_source)
            print(f"Processing page {count}")
            count += 1
            time.sleep(3)

    except Exception as ex:
        print(ex)
    finally:
        driver.close()
        driver.quit()
        print("THE END")

def get_item_url(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        src = file.read()
    # print(src)
    soup = BeautifulSoup(src, "lxml")
    item_divs = soup.find_all("div", class_ = "vertical-middle-container")

    urls = []
    for item in item_divs:
        item_url = item.find("a").get("href")
        urls.append(item_url)
    print(urls)
    print(len(urls))
    with open("item_urls.txt", "a") as file:
        for i in urls:
            file.write(f"{i}\n")

def paste_urls_into_txt():
    os.chdir("pages")
    pages_list = os.listdir()
    print(pages_list)
    for i in pages_list:
        get_item_url(i)

def get_data(file_path):
    wb = openpyxl.open('test2.xlsx')
    ws = wb.active
    with open(file_path) as file:
        url_list = file.readlines()

        url_list = [url.strip() for url in url_list]
        # print(url_list)
    result_list = []
    count = 1
    excel_count = 2
    urls_count = len(url_list)
    for url in url_list:
        response = requests.get(url=url, headers=headers)
        # print(response)
        soup = BeautifulSoup(response.text, "lxml")

        try:
            item_name = soup.find("span", {"class": "with-tooltip"}).text.strip()
        except Exception as _ex:
            item_name = None
        # print(item_name)

        try:
            item_city = soup.find("span", {"class": "locality"}).text.strip()
        except Exception as _ex:
            item_city = None
        # print(item_city)

        try:
            item_rating = soup.find("div", {"class": "rating-value text-freelancehunt bold with-tooltip"}).text.strip()
        except Exception as _ex:
            item_rating = None
        # print(item_rating)

        try:
            item_fop = soup.find("div", {"class": "col-md-6"}).find("span", {"class": "with-tooltip"}).text.strip()
        except Exception as _ex:
            item_fop = None
        # print(item_fop)

        all_list = []
        try:
            item_all = soup.find_all("div", {"class": "col-md-6"})
            try:
                for item in item_all:
                    all_list.append(str(item))
            except Exception as ex:
                print(ex)
        except Exception as _ex:
            item_all = None

        times = []
        try:
            for div in all_list:
                if "тому" in div:
                    times.append(div[(div.find("тому")-17):div.find("тому")])
            last_visit = times[0].strip()
            last_project = times[1].strip()
            # print(last_visit, last_project)
        except Exception as _ex:
            last_visit = None
            last_project = None
        try:
            item_cv = soup.find("section", {"id": "cv"}).text
        except Exception as _ex:
            item_cv = None
        # print(item_cv)
        # result_list = []
        result_list.append(
            {
                "item_name": item_name,
                "item_url": url,
                "item_city": item_city,
                "item_rating": item_rating,
                "item_fop": item_fop,
                "item_last_visit": last_visit,
                "item_last_project": last_project,
                "item_cv": item_cv
            }
        )

        # for i in range(len(result_list)):
        #     ws[]
        #     wb.save()
        try:
            ws[excel_count][0].value = item_name
            ws[excel_count][1].value = url
            ws[excel_count][2].value = item_city
            ws[excel_count][3].value = item_rating
            ws[excel_count][4].value = item_fop
            ws[excel_count][5].value = last_visit
            ws[excel_count][6].value = last_project
            ws[excel_count][7].value = item_cv

            # wb.save("test2.xlsx")
        except:
            # wb.save("test2.xlsx")
            print("exception")
        finally:
            wb.save("test2.xlsx")
            excel_count += 1
            time.sleep(random.randrange(3, 10))
            if count%10 == 0:
                time.sleep(random.randrange(5, 9))

            print(f"[+] Processed: {count}/{urls_count}")
            count += 1
        # try:
        #     with open("result.json", "a") as file:
        #         json.dump(result_list[count-2], file, indent=4, ensure_ascii=False)
        # except:
        #     print("Exception")

    wb.close()
    return "[INFO] Data collected successfully!"


def main():
    print("Starting...")
    # get_source_html('https://freelancehunt.com/ua/freelancers/programuvannya/1c')
    # paste_urls_into_txt()
    get_data(r"pages\item_urls.txt")

    # pandas.read_json("result.json").to_excel("r1.xlsx")



if __name__ == '__main__':
    main()
